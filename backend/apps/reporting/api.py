from ninja import Router
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta

router = Router()


@router.get("/statistiques", auth=None)
def statistiques(request):
    from apps.caveaux.models import Caveau, StatutCaveau
    from apps.reservations.models import Reservation
    from apps.paiements.models import Paiement

    # Statistiques caveaux
    total_caveaux = Caveau.objects.count()
    disponibles = Caveau.objects.filter(statut=StatutCaveau.DISPONIBLE).count()
    occupes = Caveau.objects.filter(statut=StatutCaveau.OCCUPE).count()
    reserves = Caveau.objects.filter(statut=StatutCaveau.RESERVE).count()
    non_exploitables = Caveau.objects.filter(statut=StatutCaveau.NON_EXPLOITABLE).count()
    taux_occupation = round((occupes / total_caveaux * 100), 1) if total_caveaux > 0 else 0

    # Statistiques réservations
    total_reservations = Reservation.objects.count()
    reservations_en_attente = Reservation.objects.filter(statut="EN_ATTENTE").count()
    reservations_validees = Reservation.objects.filter(statut="VALIDEE").count()
    reservations_refusees = Reservation.objects.filter(statut="REFUSEE").count()

    # Statistiques financières
    total_paye = Paiement.objects.filter(statut="CONFIRME").aggregate(
        total=Sum("montant_xaf")
    )["total"] or 0

    paiements_mois = Paiement.objects.filter(
        statut="CONFIRME",
        date_paiement__month=timezone.now().month,
        date_paiement__year=timezone.now().year,
    ).aggregate(total=Sum("montant_xaf"))["total"] or 0

    # Réservations par type de concession
    types_concession = Reservation.objects.filter(
        statut="VALIDEE"
    ).values("type_concession").annotate(total=Count("id"))

    # Paiements par canal
    paiements_par_canal = Paiement.objects.filter(
        statut="CONFIRME"
    ).values("canal").annotate(
        total=Count("id"),
        montant=Sum("montant_xaf")
    )
    
    # Taux d'occupation par bloc
    from apps.terrain.models import Bloc
    from apps.caveaux.models import Caveau

    blocs_stats = []
    for bloc in Bloc.objects.select_related("zone").filter(est_actif=True):
        total_bloc = Caveau.objects.filter(bloc=bloc).count()
        occupes_bloc = Caveau.objects.filter(bloc=bloc, statut="OCCUPE").count()
        reserves_bloc = Caveau.objects.filter(bloc=bloc, statut="RESERVE").count()
        disponibles_bloc = Caveau.objects.filter(bloc=bloc, statut="DISPONIBLE").count()
        taux_bloc = round((occupes_bloc / total_bloc * 100), 1) if total_bloc > 0 else 0
        blocs_stats.append({
            "bloc": str(bloc),
            "zone": bloc.zone.code,
            "total": total_bloc,
            "occupes": occupes_bloc,
            "reserves": reserves_bloc,
            "disponibles": disponibles_bloc,
            "taux_occupation": taux_bloc,
        })

    return {
        "caveaux": {
            "total": total_caveaux,
            "disponibles": disponibles,
            "occupes": occupes,
            "reserves": reserves,
            "non_exploitables": non_exploitables,
            "taux_occupation": taux_occupation,
        },
        "reservations": {
            "total": total_reservations,
            "en_attente": reservations_en_attente,
            "validees": reservations_validees,
            "refusees": reservations_refusees,
        },
        "finances": {
            "total_paye_xaf": int(total_paye),
            "paiements_mois_xaf": int(paiements_mois),
        },
        "types_concession": list(types_concession),
        "paiements_par_canal": list(paiements_par_canal),
        "blocs_stats": blocs_stats,
    }


@router.get("/export/reservations", auth=None)
def export_reservations_csv(request):
    import csv
    from django.http import HttpResponse
    from apps.reservations.models import Reservation

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="reservations.csv"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "Numéro", "Statut", "Type concession", "Montant (FCFA)",
        "Défunt", "Date décès", "Date soumission", "Date validation"
    ])

    for r in Reservation.objects.select_related("defunt").all():
        writer.writerow([
            r.numero,
            r.get_statut_display(),
            r.type_concession,
            int(r.montant_total_xaf),
            r.defunt.nom_complet,
            r.defunt.date_deces,
            r.date_soumission.strftime("%d/%m/%Y"),
            r.date_validation.strftime("%d/%m/%Y") if r.date_validation else "",
        ])

    return response


@router.get("/export/paiements", auth=None)
def export_paiements_csv(request):
    import csv
    from django.http import HttpResponse
    from apps.paiements.models import Paiement

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="paiements.csv"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "Référence", "Statut", "Canal", "Montant (FCFA)",
        "Numéro transaction", "Date paiement", "Date confirmation"
    ])

    for p in Paiement.objects.all():
        writer.writerow([
            p.reference,
            p.get_statut_display(),
            p.get_canal_display(),
            int(p.montant_xaf),
            p.numero_transaction,
            p.date_paiement.strftime("%d/%m/%Y"),
            p.date_confirmation.strftime("%d/%m/%Y") if p.date_confirmation else "",
        ])

    return response
@router.get("/export/reservations-excel", auth=None)
def export_reservations_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from apps.reservations.models import Reservation
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Réservations"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1B5E20")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Numéro", "Statut", "Type concession", "Montant (FCFA)", "Défunt", "Date décès", "Date soumission", "Date validation"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, w in enumerate([15, 12, 18, 16, 25, 14, 16, 16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    couleurs_statut = {
        "EN_ATTENTE": "FF8C00", "VALIDEE": "2E7D32",
        "REFUSEE": "C62828", "ANNULEE": "757575",
    }

    for r in Reservation.objects.select_related("defunt").all():
        ws.append([
            r.numero, r.get_statut_display(), r.type_concession,
            int(r.montant_total_xaf), r.defunt.nom_complet,
            str(r.defunt.date_deces),
            r.date_soumission.strftime("%d/%m/%Y"),
            r.date_validation.strftime("%d/%m/%Y") if r.date_validation else "",
        ])
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color=couleurs_statut.get(r.statut, "757575"))
        for cell in ws[ws.max_row]:
            cell.border = border

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reservations.xlsx"'
    return response


@router.get("/export/paiements-excel", auth=None)
def export_paiements_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from apps.paiements.models import Paiement
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paiements"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1B5E20")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Référence", "Statut", "Canal", "Montant (FCFA)", "N° Transaction", "Date paiement", "Date confirmation"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, w in enumerate([18, 12, 14, 16, 20, 16, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for p in Paiement.objects.all():
        ws.append([
            p.reference, p.get_statut_display(), p.get_canal_display(),
            int(p.montant_xaf), p.numero_transaction or "",
            p.date_paiement.strftime("%d/%m/%Y"),
            p.date_confirmation.strftime("%d/%m/%Y") if p.date_confirmation else "",
        ])
        for cell in ws[ws.max_row]:
            cell.border = border

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="paiements.xlsx"'
    return response